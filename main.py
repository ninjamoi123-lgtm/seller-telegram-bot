import os
import re
import sqlite3
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from telegram import Update, ReplyKeyboardMarkup
from telegram.ext import Application, CommandHandler, MessageHandler, ContextTypes, filters


# ========= ĞĞĞ¡Ğ¢Ğ ĞĞ™ĞšĞ˜ =========
TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN", "")
DB_PATH = os.getenv("DB_PATH", "bot.db")
TMP_DIR = Path(os.getenv("TMP_DIR", "/tmp"))

if not TELEGRAM_BOT_TOKEN:
Â  Â  raise RuntimeError("ĞĞµÑ‚ TELEGRAM_BOT_TOKEN. Ğ”Ğ¾Ğ±Ğ°Ğ²ÑŒ Ğ¿ĞµÑ€ĞµĞ¼ĞµĞ½Ğ½ÑƒÑ Ğ¾ĞºÑ€ÑƒĞ¶ĞµĞ½Ğ¸Ñ TELEGRAM_BOT_TOKEN Ğ² Railway.")

TMP_DIR.mkdir(parents=True, exist_ok=True)


# ========= ĞšĞĞĞŸĞšĞ˜ =========
MAIN_KB = ReplyKeyboardMarkup(
Â  Â  keyboard=[
Â  Â  Â  Â  ["ğŸ“ˆ ĞŸÑ€Ğ¸Ğ±Ñ‹Ğ»ÑŒ Ğ·Ğ° Ğ¿ĞµÑ€Ğ¸Ğ¾Ğ´"],
Â  Â  Â  Â  ["ğŸ“¦ Ğ—Ğ°Ğ³Ñ€ÑƒĞ·Ğ¸Ñ‚ÑŒ ÑĞµĞ±ĞµÑÑ‚Ğ¾Ğ¸Ğ¼Ğ¾ÑÑ‚ÑŒ (SKU â†’ â‚½)"],
Â  Â  Â  Â  ["â¬…ï¸ Ğ’ Ğ¼ĞµĞ½Ñ"],
Â  Â  ],
Â  Â  resize_keyboard=True
)

MODE_KB = ReplyKeyboardMarkup(
Â  Â  keyboard=[
Â  Â  Â  Â  ["ğŸŸ¡ Ğ”ĞµĞ½ÑŒĞ³Ğ¸ Ğ¾Ñ‚ OZON"],
Â  Â  Â  Â  ["ğŸŸ¢ Ğ§Ğ¸ÑÑ‚Ğ°Ñ Ğ¿Ñ€Ğ¸Ğ±Ñ‹Ğ»ÑŒ"],
Â  Â  Â  Â  ["â¬…ï¸ Ğ’ Ğ¼ĞµĞ½Ñ"],
Â  Â  ],
Â  Â  resize_keyboard=True
)

BACK_TO_MENU_KB = ReplyKeyboardMarkup(
Â  Â  keyboard=[
Â  Â  Â  Â  ["â¬…ï¸ Ğ’ Ğ¼ĞµĞ½Ñ"],
Â  Â  ],
Â  Â  resize_keyboard=True
)


# ========= SQLITE =========
def db():
Â  Â  return sqlite3.connect(DB_PATH)

def init_db():
Â  Â  with db() as conn:
Â  Â  Â  Â  conn.execute("""
Â  Â  Â  Â  CREATE TABLE IF NOT EXISTS cogs (
Â  Â  Â  Â  Â  Â  tg_id INTEGER NOT NULL,
Â  Â  Â  Â  Â  Â  sku TEXT NOT NULL,
Â  Â  Â  Â  Â  Â  cogs REAL NOT NULL,
Â  Â  Â  Â  Â  Â  updated_at TEXT NOT NULL,
Â  Â  Â  Â  Â  Â  PRIMARY KEY (tg_id, sku)
Â  Â  Â  Â  )
Â  Â  Â  Â  """)
Â  Â  Â  Â  conn.execute("""
Â  Â  Â  Â  CREATE TABLE IF NOT EXISTS profit_reports (
Â  Â  Â  Â  Â  Â  id INTEGER PRIMARY KEY AUTOINCREMENT,
Â  Â  Â  Â  Â  Â  tg_id INTEGER NOT NULL,
Â  Â  Â  Â  Â  Â  mode TEXT NOT NULL,
Â  Â  Â  Â  Â  Â  file_name TEXT NOT NULL,
Â  Â  Â  Â  Â  Â  revenue REAL NOT NULL,
Â  Â  Â  Â  Â  Â  deductions REAL NOT NULL,
Â  Â  Â  Â  Â  Â  net_mp REAL NOT NULL,
Â  Â  Â  Â  Â  Â  cogs_total REAL,
Â  Â  Â  Â  Â  Â  net_profit REAL,
Â  Â  Â  Â  Â  Â  margin REAL,
Â  Â  Â  Â  Â  Â  created_at TEXT NOT NULL,
Â  Â  Â  Â  Â  Â  note TEXT
Â  Â  Â  Â  )
Â  Â  Â  Â  """)

def upsert_cogs(tg_id: int, sku: str, cogs_val: float):
Â  Â  with db() as conn:
Â  Â  Â  Â  conn.execute(
Â  Â  Â  Â  Â  Â  "INSERT INTO cogs(tg_id, sku, cogs, updated_at) VALUES(?,?,?,?) "
Â  Â  Â  Â  Â  Â  "ON CONFLICT(tg_id, sku) DO UPDATE SET cogs=excluded.cogs, updated_at=excluded.updated_at",
Â  Â  Â  Â  Â  Â  (tg_id, sku, float(cogs_val), datetime.utcnow().strftime("%Y-%m-%d"))
Â  Â  Â  Â  )

def get_cogs_map(tg_id: int) -> dict:
Â  Â  with db() as conn:
Â  Â  Â  Â  rows = conn.execute("SELECT sku, cogs FROM cogs WHERE tg_id=?", (tg_id,)).fetchall()
Â  Â  return {r[0]: float(r[1]) for r in rows}

def save_report(tg_id: int, payload: dict):
Â  Â  with db() as conn:
Â  Â  Â  Â  conn.execute("""
Â  Â  Â  Â  INSERT INTO profit_reports(
Â  Â  Â  Â  Â  Â  tg_id, mode, file_name, revenue, deductions, net_mp,
Â  Â  Â  Â  Â  Â  cogs_total, net_profit, margin, created_at, note
Â  Â  Â  Â  ) VALUES (?,?,?,?,?,?,?,?,?,?,?)
Â  Â  Â  Â  """, (
Â  Â  Â  Â  Â  Â  tg_id,
Â  Â  Â  Â  Â  Â  payload["mode"],
Â  Â  Â  Â  Â  Â  payload["file_name"],
Â  Â  Â  Â  Â  Â  float(payload["revenue"]),
Â  Â  Â  Â  Â  Â  float(payload["deductions"]),
Â  Â  Â  Â  Â  Â  float(payload["net_mp"]),
Â  Â  Â  Â  Â  Â  payload.get("cogs_total"),
Â  Â  Â  Â  Â  Â  payload.get("net_profit"),
Â  Â  Â  Â  Â  Â  payload.get("margin"),
Â  Â  Â  Â  Â  Â  datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"),
Â  Â  Â  Â  Â  Â  payload.get("note", ""),
Â  Â  Â  Â  ))


# ========= Ğ£Ğ¢Ğ˜Ğ›Ğ˜Ğ¢Ğ« =========
def money(x: float) -> str:
Â  Â  if x is None:
Â  Â  Â  Â  return "0 â‚½"
Â  Â  if abs(x - int(x)) < 1e-9:
Â  Â  Â  Â  return f"{int(x)} â‚½"
Â  Â  return f"{x:.2f} â‚½"

def pct(x: float) -> str:
Â  Â  return f"{x:.2f}%"

def norm(s: str) -> str:
Â  Â  return re.sub(r"\s+", " ", str(s)).strip().lower()

def parse_number(x):
Â  Â  if x is None:
Â  Â  Â  Â  return None
Â  Â  s = str(x).replace("\u00A0", "").replace(" ", "").replace(",", ".").strip()
Â  Â  m = re.search(r"-?\d+(?:\.\d+)?", s)
Â  Â  if not m:
Â  Â  Â  Â  return None
Â  Â  try:
Â  Â  Â  Â  return float(m.group(0))
Â  Â  except Exception:
Â  Â  Â  Â  return None


# ========= Ğ§Ğ¢Ğ•ĞĞ˜Ğ• XLSX =========
def load_xlsx_rows(path: str):
Â  Â  wb = load_workbook(path, data_only=True)
Â  Â  ws = wb.active
Â  Â  rows = list(ws.iter_rows(values_only=True))
Â  Â  if not rows or len(rows) < 2:
Â  Â  Â  Â  raise ValueError("Ğ¤Ğ°Ğ¹Ğ» Ğ¿ÑƒÑÑ‚Ğ¾Ğ¹ Ğ¸Ğ»Ğ¸ Ğ±ĞµĞ· Ğ´Ğ°Ğ½Ğ½Ñ‹Ñ….")
Â  Â  header = [str(h).strip() if h is not None else "" for h in rows[0]]
Â  Â  data = rows[1:]
Â  Â  return header, data

def find_col_index(header, keywords):
Â  Â  for i, col in enumerate(header):
Â  Â  Â  Â  nc = norm(col)
Â  Â  Â  Â  for kw in keywords:
Â  Â  Â  Â  Â  Â  if kw in nc:
Â  Â  Â  Â  Â  Â  Â  Â  return i
Â  Â  return None

def parse_report_xlsx(path: str):
Â  Â  header, data = load_xlsx_rows(path)

Â  Â  amount_idx = find_col_index(header, ["Ğ¸Ñ‚Ğ¾Ğ³Ğ¾"])
Â  Â  if amount_idx is None:
Â  Â  Â  Â  amount_idx = find_col_index(header, ["ÑÑƒĞ¼Ğ¼", "Ğ½Ğ°Ñ‡Ğ¸ÑĞ»", "amount"])

Â  Â  if amount_idx is None:
Â  Â  Â  Â  # Ğ·Ğ°Ğ¿Ğ°ÑĞ½Ğ¾Ğ¹ Ğ²Ğ°Ñ€Ğ¸Ğ°Ğ½Ñ‚: Ğ¸Ñ‰ĞµĞ¼ ĞºĞ¾Ğ»Ğ¾Ğ½ĞºÑƒ Ñ Ğ½Ğ°Ğ¸Ğ±Ğ¾Ğ»ÑŒÑˆĞ¸Ğ¼ ĞºĞ¾Ğ»Ğ¸Ñ‡ĞµÑÑ‚Ğ²Ğ¾Ğ¼ Ñ‡Ğ¸ÑĞµĞ»
Â  Â  Â  Â  best_i, best_score = None, 0
Â  Â  Â  Â  for i in range(len(header)):
Â  Â  Â  Â  Â  Â  score = 0
Â  Â  Â  Â  Â  Â  for r in data[:2000]:
Â  Â  Â  Â  Â  Â  Â  Â  v = parse_number(r[i] if i < len(r) else None)
Â  Â  Â  Â  Â  Â  Â  Â  if v is not None:
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  score += 1
Â  Â  Â  Â  Â  Â  if score > best_score:
Â  Â  Â  Â  Â  Â  Â  Â  best_score = score
Â  Â  Â  Â  Â  Â  Â  Â  best_i = i
Â  Â  Â  Â  amount_idx = best_i

Â  Â  if amount_idx is None:
Â  Â  Â  Â  raise ValueError("ĞĞµ Ğ½Ğ°ÑˆÑ‘Ğ» ĞºĞ¾Ğ»Ğ¾Ğ½ĞºÑƒ Ñ ÑÑƒĞ¼Ğ¼Ğ¾Ğ¹/Ğ¸Ñ‚Ğ¾Ğ³Ğ¾.")

Â  Â  sku_idx = find_col_index(header, ["sku", "offer", "Ğ°Ñ€Ñ‚Ğ¸ĞºÑƒĞ»"])
Â  Â  qty_idx = find_col_index(header, ["ĞºĞ¾Ğ»", "quantity", "qty"])

Â  Â  revenue = 0.0
Â  Â  deductions = 0.0
Â  Â  total = 0.0

Â  Â  by_sku_amount = {}

Â  Â  for r in data:
Â  Â  Â  Â  if amount_idx >= len(r):
Â  Â  Â  Â  Â  Â  continue
Â  Â  Â  Â  amt = parse_number(r[amount_idx])
Â  Â  Â  Â  if amt is None:
Â  Â  Â  Â  Â  Â  continue

Â  Â  Â  Â  total += amt
Â  Â  Â  Â  if amt > 0:
Â  Â  Â  Â  Â  Â  revenue += amt
Â  Â  Â  Â  elif amt < 0:
Â  Â  Â  Â  Â  Â  deductions += amt

Â  Â  Â  Â  sku = ""
Â  Â  Â  Â  if sku_idx is not None and sku_idx < len(r):
Â  Â  Â  Â  Â  Â  sku = str(r[sku_idx]).strip() if r[sku_idx] is not None else ""

Â  Â  Â  Â  if sku:
Â  Â  Â  Â  Â  Â  by_sku_amount[sku] = by_sku_amount.get(sku, 0.0) + amt

Â  Â  note = f"amount_col_idx={amount_idx} | sku_idx={sku_idx if sku_idx is not None else 'NOT_FOUND'} | qty_idx={qty_idx if qty_idx is not None else 'NOT_FOUND'}"
Â  Â  return {
Â  Â  Â  Â  "revenue": float(revenue),
Â  Â  Â  Â  "deductions": float(deductions),
Â  Â  Â  Â  "total": float(total),
Â  Â  Â  Â  "by_sku_amount": by_sku_amount,
Â  Â  Â  Â  "note": note,
Â  Â  Â  Â  "header": header,
Â  Â  Â  Â  "amount_idx": amount_idx,
Â  Â  Â  Â  "sku_idx": sku_idx,
Â  Â  Â  Â  "qty_idx": qty_idx,
Â  Â  Â  Â  "data": data,
Â  Â  }

def top_lines_dict(d: dict, n=5, ascending=False):
Â  Â  if not d:
Â  Â  Â  Â  return "Ğ½ĞµÑ‚ Ğ´Ğ°Ğ½Ğ½Ñ‹Ñ…"
Â  Â  items = sorted(d.items(), key=lambda x: x[1], reverse=not ascending)[:n]
Â  Â  return "\n".join([f"{k} â€” {money(float(v))}" for k, v in items])

def parse_cogs_xlsx(path: str):
Â  Â  header, data = load_xlsx_rows(path)

Â  Â  sku_idx = find_col_index(header, ["sku", "Ğ°Ñ€Ñ‚Ğ¸ĞºÑƒĞ»", "offer"])
Â  Â  cogs_idx = find_col_index(header, ["cogs", "ÑĞµĞ±ĞµÑÑ‚", "ÑĞµĞ±ĞµÑÑ‚Ğ¾Ğ¸Ğ¼", "cost"])

Â  Â  if sku_idx is None or cogs_idx is None:
Â  Â  Â  Â  raise ValueError("Ğ’ Ñ„Ğ°Ğ¹Ğ»Ğµ ÑĞµĞ±ĞµÑÑ‚Ğ¾Ğ¸Ğ¼Ğ¾ÑÑ‚Ğ¸ Ğ½ÑƒĞ¶Ğ½Ñ‹ ĞºĞ¾Ğ»Ğ¾Ğ½ĞºĞ¸: sku Ğ¸ cogs (Ğ¸Ğ»Ğ¸ 'Ğ°Ñ€Ñ‚Ğ¸ĞºÑƒĞ»' Ğ¸ 'ÑĞµĞ±ĞµÑÑ‚Ğ¾Ğ¸Ğ¼Ğ¾ÑÑ‚ÑŒ').")

Â  Â  rows = []
Â  Â  for r in data:
Â  Â  Â  Â  if sku_idx >= len(r) or cogs_idx >= len(r):
Â  Â  Â  Â  Â  Â  continue
Â  Â  Â  Â  sku = str(r[sku_idx]).strip() if r[sku_idx] is not None else ""
Â  Â  Â  Â  cogs_val = parse_number(r[cogs_idx])
Â  Â  Â  Â  if not sku or cogs_val is None:
Â  Â  Â  Â  Â  Â  continue
Â  Â  Â  Â  rows.append((sku, float(cogs_val)))
Â  Â  return rows


# ========= BOT =========
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
Â  Â  context.user_data.clear()
Â  Â  await update.message.reply_text(
Â  Â  Â  Â  "ĞŸÑ€Ğ¸Ğ²ĞµÑ‚ ğŸ‘‹\n\n"
Â  Â  Â  Â  "Ğ¯ ÑÑ‡Ğ¸Ñ‚Ğ°Ñ Ğ¿Ñ€Ğ¸Ğ±Ñ‹Ğ»ÑŒ Ğ¿Ğ¾ Ğ¾Ñ‚Ñ‡Ñ‘Ñ‚Ñƒ OZON Â«ĞĞ°Ñ‡Ğ¸ÑĞ»ĞµĞ½Ğ¸ÑÂ».\n"
Â  Â  Â  Â  "âœ… Ğ¤Ğ¾Ñ€Ğ¼Ğ°Ñ‚ Ñ„Ğ°Ğ¹Ğ»Ğ¾Ğ²: Ğ¢ĞĞ›Ğ¬ĞšĞ Excel (.xlsx)\n\n"
Â  Â  Â  Â  "Ğ’Ñ‹Ğ±ĞµÑ€Ğ¸ Ğ´ĞµĞ¹ÑÑ‚Ğ²Ğ¸Ğµ â¬‡ï¸",
Â  Â  Â  Â  reply_markup=MAIN_KB
Â  Â  )

async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
Â  Â  text = (update.message.text or "").strip()
Â  Â  print("TEXT:", repr(text), "STATE:", dict(context.user_data))

Â  Â  # ĞºĞ½Ğ¾Ğ¿ĞºĞ¸ â€” Ğ² Ğ½Ğ°Ñ‡Ğ°Ğ»Ğµ
Â  Â  if text == "â¬…ï¸ Ğ’ Ğ¼ĞµĞ½Ñ":
Â  Â  Â  Â  context.user_data.clear()
Â  Â  Â  Â  await update.message.reply_text("ĞœĞµĞ½Ñ â¬‡ï¸", reply_markup=MAIN_KB)
Â  Â  Â  Â  return

Â  Â  if text == "ğŸ“ˆ ĞŸÑ€Ğ¸Ğ±Ñ‹Ğ»ÑŒ Ğ·Ğ° Ğ¿ĞµÑ€Ğ¸Ğ¾Ğ´":
Â  Â  Â  Â  context.user_data.clear()
Â  Â  Â  Â  await update.message.reply_text(
Â  Â  Â  Â  Â  Â  "ĞšĞ°Ğº ÑÑ‡Ğ¸Ñ‚Ğ°Ñ‚ÑŒ?\n\n"
Â  Â  Â  Â  Â  Â  "ğŸŸ¡ Ğ”ĞµĞ½ÑŒĞ³Ğ¸ Ğ¾Ñ‚ OZON â€” Ğ¸Ñ‚Ğ¾Ğ³ Ğ¿Ğ¾ Ğ¾Ñ‚Ñ‡Ñ‘Ñ‚Ñƒ (Ğ² Ğ¿Ğ»ÑÑĞµ/Ğ² Ğ¼Ğ¸Ğ½ÑƒÑĞµ)\n"
Â  Â  Â  Â  Â  Â  "ğŸŸ¢ Ğ§Ğ¸ÑÑ‚Ğ°Ñ Ğ¿Ñ€Ğ¸Ğ±Ñ‹Ğ»ÑŒ â€” Ğ½ÑƒĞ¶Ğ½Ğ° ÑĞµĞ±ĞµÑÑ‚Ğ¾Ğ¸Ğ¼Ğ¾ÑÑ‚ÑŒ SKU â†’ â‚½\n\n"
Â  Â  Â  Â  Â  Â  "ĞŸÑ€Ğ¸ÑˆĞ»Ğ¸ .xlsx Ñ„Ğ°Ğ¹Ğ» Ğ¿Ğ¾ÑĞ»Ğµ Ğ²Ñ‹Ğ±Ğ¾Ñ€Ğ° Ñ€ĞµĞ¶Ğ¸Ğ¼Ğ° â¬‡ï¸",
Â  Â  Â  Â  Â  Â  reply_markup=MODE_KB
Â  Â  Â  Â  )
Â  Â  Â  Â  return

Â  Â  if text == "ğŸŸ¡ Ğ”ĞµĞ½ÑŒĞ³Ğ¸ Ğ¾Ñ‚ OZON":
Â  Â  Â  Â  context.user_data.clear()
Â  Â  Â  Â  context.user_data["mode"] = "mp_money"
Â  Â  Â  Â  context.user_data["await_report"] = True
Â  Â  Â  Â  await update.message.reply_text(
Â  Â  Â  Â  Â  Â  "ĞĞº. ĞŸÑ€Ğ¸ÑˆĞ»Ğ¸ .xlsx Ğ¾Ñ‚Ñ‡Ñ‘Ñ‚ OZON Â«ĞĞ°Ñ‡Ğ¸ÑĞ»ĞµĞ½Ğ¸ÑÂ» Ğ·Ğ° Ğ½ÑƒĞ¶Ğ½Ñ‹Ğ¹ Ğ¿ĞµÑ€Ğ¸Ğ¾Ğ´.",
Â  Â  Â  Â  Â  Â  reply_markup=MODE_KB
Â  Â  Â  Â  )
Â  Â  Â  Â  return

Â  Â  if text == "ğŸŸ¢ Ğ§Ğ¸ÑÑ‚Ğ°Ñ Ğ¿Ñ€Ğ¸Ğ±Ñ‹Ğ»ÑŒ":
Â  Â  Â  Â  context.user_data.clear()
Â  Â  Â  Â  context.user_data["mode"] = "net_profit"
Â  Â  Â  Â  context.user_data["await_report"] = True
Â  Â  Â  Â  await update.message.reply_text(
Â  Â  Â  Â  Â  Â  "ĞĞº. ĞŸÑ€Ğ¸ÑˆĞ»Ğ¸ .xlsx Ğ¾Ñ‚Ñ‡Ñ‘Ñ‚ OZON Â«ĞĞ°Ñ‡Ğ¸ÑĞ»ĞµĞ½Ğ¸ÑÂ» Ğ·Ğ° Ğ½ÑƒĞ¶Ğ½Ñ‹Ğ¹ Ğ¿ĞµÑ€Ğ¸Ğ¾Ğ´.\n\n"
Â  Â  Â  Â  Â  Â  "Ğ•ÑĞ»Ğ¸ ÑĞµĞ±ĞµÑÑ‚Ğ¾Ğ¸Ğ¼Ğ¾ÑÑ‚ÑŒ ĞµÑ‰Ñ‘ Ğ½Ğµ Ğ·Ğ°Ğ³Ñ€ÑƒĞ¶Ğ°Ğ» â€” ÑĞ½Ğ°Ñ‡Ğ°Ğ»Ğ° Ğ·Ğ°Ğ³Ñ€ÑƒĞ·Ğ¸ Ñ‡ĞµÑ€ĞµĞ· Â«ğŸ“¦ Ğ—Ğ°Ğ³Ñ€ÑƒĞ·Ğ¸Ñ‚ÑŒ ÑĞµĞ±ĞµÑÑ‚Ğ¾Ğ¸Ğ¼Ğ¾ÑÑ‚ÑŒ (SKU â†’ â‚½)Â».",
Â  Â  Â  Â  Â  Â  reply_markup=MODE_KB
Â  Â  Â  Â  )
Â  Â  Â  Â  return

Â  Â  if text == "ğŸ“¦ Ğ—Ğ°Ğ³Ñ€ÑƒĞ·Ğ¸Ñ‚ÑŒ ÑĞµĞ±ĞµÑÑ‚Ğ¾Ğ¸Ğ¼Ğ¾ÑÑ‚ÑŒ (SKU â†’ â‚½)":
Â  Â  Â  Â  context.user_data.clear()
Â  Â  Â  Â  context.user_data["await_cogs"] = True
Â  Â  Â  Â  await update.message.reply_text(
Â  Â  Â  Â  Â  Â  "ĞŸÑ€Ğ¸ÑˆĞ»Ğ¸ Excel (.xlsx) Ñ„Ğ°Ğ¹Ğ» ÑĞµĞ±ĞµÑÑ‚Ğ¾Ğ¸Ğ¼Ğ¾ÑÑ‚Ğ¸.\n\n"
Â  Â  Â  Â  Â  Â  "Ğ’ Ñ‚Ğ°Ğ±Ğ»Ğ¸Ñ†Ğµ Ğ´Ğ¾Ğ»Ğ¶Ğ½Ñ‹ Ğ±Ñ‹Ñ‚ÑŒ ĞºĞ¾Ğ»Ğ¾Ğ½ĞºĞ¸:\n"
Â  Â  Â  Â  Â  Â  "â€¢ sku (Ğ¸Ğ»Ğ¸ ĞÑ€Ñ‚Ğ¸ĞºÑƒĞ»/offer)\n"
Â  Â  Â  Â  Â  Â  "â€¢ cogs (Ğ¸Ğ»Ğ¸ Ğ¡ĞµĞ±ĞµÑÑ‚Ğ¾Ğ¸Ğ¼Ğ¾ÑÑ‚ÑŒ)\n\n"
Â  Â  Â  Â  Â  Â  "ĞŸÑ€Ğ¸Ğ¼ĞµÑ€ Ğ·Ğ°Ğ³Ğ¾Ğ»Ğ¾Ğ²ĞºĞ¾Ğ²: sku | cogs",
Â  Â  Â  Â  Â  Â  reply_markup=BACK_TO_MENU_KB
Â  Â  Â  Â  )
Â  Â  Â  Â  return

Â  Â  # ĞµÑĞ»Ğ¸ Ğ¶Ğ´Ñ‘Ğ¼ Ñ„Ğ°Ğ¹Ğ»
Â  Â  if context.user_data.get("await_report"):
Â  Â  Â  Â  await update.message.reply_text("Ğ¯ Ğ¶Ğ´Ñƒ .xlsx Ğ¾Ñ‚Ñ‡Ñ‘Ñ‚ Â«ĞĞ°Ñ‡Ğ¸ÑĞ»ĞµĞ½Ğ¸ÑÂ». ĞŸÑ€Ğ¸ÑˆĞ»Ğ¸ Ğ´Ğ¾ĞºÑƒĞ¼ĞµĞ½Ñ‚Ğ¾Ğ¼.", reply_markup=MODE_KB)
Â  Â  Â  Â  return

Â  Â  if context.user_data.get("await_cogs"):
Â  Â  Â  Â  await update.message.reply_text("Ğ¯ Ğ¶Ğ´Ñƒ .xlsx Ñ„Ğ°Ğ¹Ğ» ÑĞµĞ±ĞµÑÑ‚Ğ¾Ğ¸Ğ¼Ğ¾ÑÑ‚Ğ¸. ĞŸÑ€Ğ¸ÑˆĞ»Ğ¸ Ğ´Ğ¾ĞºÑƒĞ¼ĞµĞ½Ñ‚Ğ¾Ğ¼.", reply_markup=BACK_TO_MENU_KB)
Â  Â  Â  Â  return

Â  Â  await update.message.reply_text("Ğ’Ñ‹Ğ±ĞµÑ€Ğ¸ Ğ´ĞµĞ¹ÑÑ‚Ğ²Ğ¸Ğµ ĞºĞ½Ğ¾Ğ¿ĞºĞ¾Ğ¹ â¬‡ï¸", reply_markup=MAIN_KB)

async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
Â  Â  doc = update.message.document
Â  Â  tg_id = update.effective_user.id
Â  Â  if not doc:
Â  Â  Â  Â  return

Â  Â  file_name = doc.file_name or "file"
Â  Â  suffix = Path(file_name).suffix.lower()
Â  Â  tg_file = await context.bot.get_file(doc.file_id)

Â  Â  # Ğ¢ĞĞ›Ğ¬ĞšĞ XLSX
Â  Â  if suffix != ".xlsx":
Â  Â  Â  Â  await update.message.reply_text("Ğ¯ Ğ¿Ñ€Ğ¸Ğ½Ğ¸Ğ¼Ğ°Ñ Ñ‚Ğ¾Ğ»ÑŒĞºĞ¾ Excel (.xlsx).", reply_markup=MAIN_KB)
Â  Â  Â  Â  return

Â  Â  # --- ÑĞµĞ±ĞµÑÑ‚Ğ¾Ğ¸Ğ¼Ğ¾ÑÑ‚ÑŒ ---
Â  Â  if context.user_data.get("await_cogs"):
Â  Â  Â  Â  local_path = str(TMP_DIR / f"cogs_{tg_id}_{int(datetime.utcnow().timestamp())}.xlsx")
Â  Â  Â  Â  await tg_file.download_to_drive(custom_path=local_path)

Â  Â  Â  Â  try:
Â  Â  Â  Â  Â  Â  rows = parse_cogs_xlsx(local_path)
Â  Â  Â  Â  Â  Â  count = 0
Â  Â  Â  Â  Â  Â  for sku, cogs_val in rows:
Â  Â  Â  Â  Â  Â  Â  Â  upsert_cogs(tg_id, sku, cogs_val)
Â  Â  Â  Â  Â  Â  Â  Â  count += 1

Â  Â  Â  Â  Â  Â  context.user_data.clear()
Â  Â  Â  Â  Â  Â  await update.message.reply_text(f"âœ… Ğ—Ğ°Ğ³Ñ€ÑƒĞ¶ĞµĞ½Ğ¾ ÑĞµĞ±ĞµÑÑ‚Ğ¾Ğ¸Ğ¼Ğ¾ÑÑ‚ĞµĞ¹: {count} SKU", reply_markup=MAIN_KB)
Â  Â  Â  Â  except Exception as e:
Â  Â  Â  Â  Â  Â  await update.message.reply_text(f"ĞÑˆĞ¸Ğ±ĞºĞ° Ñ„Ğ°Ğ¹Ğ»Ğ° ÑĞµĞ±ĞµÑÑ‚Ğ¾Ğ¸Ğ¼Ğ¾ÑÑ‚Ğ¸: {e}", reply_markup=BACK_TO_MENU_KB)
Â  Â  Â  Â  return

Â  Â  # --- Ğ¾Ñ‚Ñ‡Ñ‘Ñ‚ Ğ½Ğ°Ñ‡Ğ¸ÑĞ»ĞµĞ½Ğ¸Ñ ---
Â  Â  if context.user_data.get("await_report"):
Â  Â  Â  Â  local_path = str(TMP_DIR / f"report_{tg_id}_{int(datetime.utcnow().timestamp())}.xlsx")
Â  Â  Â  Â  await tg_file.download_to_drive(custom_path=local_path)

Â  Â  Â  Â  mode = context.user_data.get("mode", "mp_money")

Â  Â  Â  Â  try:
Â  Â  Â  Â  Â  Â  parsed = parse_report_xlsx(local_path)
Â  Â  Â  Â  Â  Â  revenue = parsed["revenue"]
Â  Â  Â  Â  Â  Â  deductions = parsed["deductions"]
Â  Â  Â  Â  Â  Â  net_mp = parsed["total"]

Â  Â  Â  Â  Â  Â  if mode == "mp_money":
Â  Â  Â  Â  Â  Â  Â  Â  status = "ğŸŸ¢" if net_mp > 0 else "ğŸ”´"
Â  Â  Â  Â  Â  Â  Â  Â  msg = (
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  "ğŸ“ˆ Ğ˜Ñ‚Ğ¾Ğ³Ğ¸ Ğ¿Ğ¾ Ğ¾Ñ‚Ñ‡Ñ‘Ñ‚Ñƒ OZON Â«ĞĞ°Ñ‡Ğ¸ÑĞ»ĞµĞ½Ğ¸ÑÂ»\n\n"
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  f"ĞĞ°Ñ‡Ğ¸ÑĞ»ĞµĞ½Ğ¾ (Ğ¿Ğ»ÑÑ): {money(revenue)}\n"
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  f"Ğ£Ğ´ĞµÑ€Ğ¶Ğ°Ğ½Ğ¸Ñ (Ğ¼Ğ¸Ğ½ÑƒÑ): {money(deductions)}\n"
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  f"Ğ˜Ñ‚Ğ¾Ğ³Ğ¾ Ğ¾Ñ‚ OZON: {money(net_mp)}\n"
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  f"Ğ¡Ñ‚Ğ°Ñ‚ÑƒÑ: {status}\n\n"
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  f"Ğ¢ĞµÑ….Ğ¸Ğ½Ñ„Ğ¾: {parsed['note']}\n\n"
Â  Â  Â  Â  Â  Â  Â  Â  )

Â  Â  Â  Â  Â  Â  Â  Â  by_sku = parsed["by_sku_amount"]
Â  Â  Â  Â  Â  Â  Â  Â  if by_sku:
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  msg += "Ğ¢ĞĞŸ-5 SKU Ğ¿Ğ¾ Ğ¸Ñ‚Ğ¾Ğ³Ñƒ:\n" + top_lines_dict(by_sku, 5, ascending=False) + "\n\n"
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  msg += "Ğ¢ĞĞŸ-5 SKU Ğ² Ğ¼Ğ¸Ğ½ÑƒÑ:\n" + top_lines_dict(by_sku, 5, ascending=True) + "\n"
Â  Â  Â  Â  Â  Â  Â  Â  else:
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  msg += "Ğ¢ĞĞŸ SKU: Ğ½ĞµÑ‚ (Ğ½Ğµ Ğ½Ğ°ÑˆÑ‘Ğ» ĞºĞ¾Ğ»Ğ¾Ğ½ĞºÑƒ SKU/offer_id/Ğ°Ñ€Ñ‚Ğ¸ĞºÑƒĞ»)\n"

Â  Â  Â  Â  Â  Â  Â  Â  save_report(tg_id, {
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  "mode": "mp_money",
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  "file_name": file_name,
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  "revenue": revenue,
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  "deductions": deductions,
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  "net_mp": net_mp,
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  "note": parsed["note"],
Â  Â  Â  Â  Â  Â  Â  Â  })

Â  Â  Â  Â  Â  Â  Â  Â  context.user_data.clear()
Â  Â  Â  Â  Â  Â  Â  Â  await update.message.reply_text(msg, reply_markup=MAIN_KB)
Â  Â  Â  Â  Â  Â  Â  Â  return

Â  Â  Â  Â  Â  Â  # net_profit
Â  Â  Â  Â  Â  Â  cogs_map = get_cogs_map(tg_id)
Â  Â  Â  Â  Â  Â  if not cogs_map:
Â  Â  Â  Â  Â  Â  Â  Â  status = "ğŸŸ¢" if net_mp > 0 else "ğŸ”´"
Â  Â  Â  Â  Â  Â  Â  Â  msg = (
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  "ğŸŸ¢ Ğ§Ğ¸ÑÑ‚Ğ°Ñ Ğ¿Ñ€Ğ¸Ğ±Ñ‹Ğ»ÑŒ\n\n"
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  "Ğ¡ĞµĞ±ĞµÑÑ‚Ğ¾Ğ¸Ğ¼Ğ¾ÑÑ‚ÑŒ Ğ½Ğµ Ğ·Ğ°Ğ³Ñ€ÑƒĞ¶ĞµĞ½Ğ°.\n"
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  "ĞŸĞ¾ĞºĞ°Ğ·Ñ‹Ğ²Ğ°Ñ Ğ´ĞµĞ½ÑŒĞ³Ğ¸ Ğ¾Ñ‚ OZON (Ğ±ĞµĞ· ÑĞµĞ±ĞµÑÑ‚Ğ¾Ğ¸Ğ¼Ğ¾ÑÑ‚Ğ¸):\n\n"
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  f"Ğ˜Ñ‚Ğ¾Ğ³Ğ¾ Ğ¾Ñ‚ OZON: {money(net_mp)}\n"
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  f"Ğ¡Ñ‚Ğ°Ñ‚ÑƒÑ: {status}\n\n"
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  "Ğ—Ğ°Ğ³Ñ€ÑƒĞ·Ğ¸ ÑĞµĞ±ĞµÑÑ‚Ğ¾Ğ¸Ğ¼Ğ¾ÑÑ‚ÑŒ Ñ‡ĞµÑ€ĞµĞ· Â«ğŸ“¦ Ğ—Ğ°Ğ³Ñ€ÑƒĞ·Ğ¸Ñ‚ÑŒ ÑĞµĞ±ĞµÑÑ‚Ğ¾Ğ¸Ğ¼Ğ¾ÑÑ‚ÑŒ (SKU â†’ â‚½)Â» Ğ¸ Ğ¿Ğ¾Ğ²Ñ‚Ğ¾Ñ€Ğ¸ Ñ€Ğ°ÑÑ‡Ñ‘Ñ‚."
Â  Â  Â  Â  Â  Â  Â  Â  )
Â  Â  Â  Â  Â  Â  Â  Â  save_report(tg_id, {
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  "mode": "net_profit",
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  "file_name": file_name,
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  "revenue": revenue,
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  "deductions": deductions,
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  "net_mp": net_mp,
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  "note": "NO_COGS | " + parsed["note"],
Â  Â  Â  Â  Â  Â  Â  Â  })
Â  Â  Â  Â  Â  Â  Â  Â  context.user_data.clear()
Â  Â  Â  Â  Â  Â  Â  Â  await update.message.reply_text(msg, reply_markup=MAIN_KB)
Â  Â  Â  Â  Â  Â  Â  Â  return

Â  Â  Â  Â  Â  Â  # ÑÑ‡Ğ¸Ñ‚Ğ°ĞµĞ¼ ÑĞµĞ±ĞµÑÑ‚Ğ¾Ğ¸Ğ¼Ğ¾ÑÑ‚ÑŒ Ğ¿Ğ¾ SKU (qty Ğ¿Ğ¾ĞºĞ° Ğ½Ğµ Ğ¸ÑĞ¿Ğ¾Ğ»ÑŒĞ·ÑƒĞµĞ¼, ĞµÑĞ»Ğ¸ Ğ½ĞµÑ‚ ĞºĞ¾Ğ»Ğ¾Ğ½ĞºĞ¸ qty â€” ÑÑ‡Ğ¸Ñ‚Ğ°ĞµĞ¼ 1)
Â  Â  Â  Â  Â  Â  sku_idx = parsed["sku_idx"]
Â  Â  Â  Â  Â  Â  qty_idx = parsed["qty_idx"]
Â  Â  Â  Â  Â  Â  data = parsed["data"]

Â  Â  Â  Â  Â  Â  if sku_idx is None:
Â  Â  Â  Â  Â  Â  Â  Â  msg = (
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  "ğŸŸ¢ Ğ§Ğ¸ÑÑ‚Ğ°Ñ Ğ¿Ñ€Ğ¸Ğ±Ñ‹Ğ»ÑŒ\n\n"
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  "Ğ’ Ğ¾Ñ‚Ñ‡Ñ‘Ñ‚Ğµ Ğ½Ğµ Ğ½Ğ°ÑˆÑ‘Ğ» SKU/offer_id/Ğ°Ñ€Ñ‚Ğ¸ĞºÑƒĞ» â€” Ğ½Ğµ Ğ¼Ğ¾Ğ³Ñƒ Ğ¿Ñ€Ğ¸Ğ¼ĞµĞ½Ğ¸Ñ‚ÑŒ ÑĞµĞ±ĞµÑÑ‚Ğ¾Ğ¸Ğ¼Ğ¾ÑÑ‚ÑŒ.\n"
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  "ĞŸÑ€Ğ¾Ğ²ĞµÑ€ÑŒ, Ñ‡Ñ‚Ğ¾ Ğ² Ğ²Ñ‹Ğ³Ñ€ÑƒĞ·ĞºĞµ ĞµÑÑ‚ÑŒ Ğ°Ñ€Ñ‚Ğ¸ĞºÑƒĞ»Ñ‹."
Â  Â  Â  Â  Â  Â  Â  Â  )
Â  Â  Â  Â  Â  Â  Â  Â  context.user_data.clear()
Â  Â  Â  Â  Â  Â  Â  Â  await update.message.reply_text(msg, reply_markup=MAIN_KB)
Â  Â  Â  Â  Â  Â  Â  Â  return

Â  Â  Â  Â  Â  Â  cogs_total = 0.0
Â  Â  Â  Â  Â  Â  by_sku_profit = {}

Â  Â  Â  Â  Â  Â  for r in data:
Â  Â  Â  Â  Â  Â  Â  Â  if sku_idx >= len(r):
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  continue
Â  Â  Â  Â  Â  Â  Â  Â  sku = str(r[sku_idx]).strip() if r[sku_idx] is not None else ""
Â  Â  Â  Â  Â  Â  Â  Â  if not sku:
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  continue

Â  Â  Â  Â  Â  Â  Â  Â  amt = None
Â  Â  Â  Â  Â  Â  Â  Â  if parsed["amount_idx"] < len(r):
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  amt = parse_number(r[parsed["amount_idx"]])
Â  Â  Â  Â  Â  Â  Â  Â  if amt is None:
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  continue

Â  Â  Â  Â  Â  Â  Â  Â  qty = 1.0
Â  Â  Â  Â  Â  Â  Â  Â  if qty_idx is not None and qty_idx < len(r):
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  q = parse_number(r[qty_idx])
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  if q is not None and q > 0:
Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  Â  qty = float(q)

Â  Â  Â  Â  Â  Â  Â  Â  c = float(cogs_map.get(sku, 0.0)) * qty
Â  Â  Â  Â  Â  Â  Â  Â  cogs_total += c

Â  Â  Â  Â  Â  Â  Â  Â  # Ğ¿Ñ€Ğ¸Ğ±Ñ‹Ğ»ÑŒ Ğ¿Ğ¾ SKU = ÑÑƒĞ¼Ğ¼Ğ° Ğ½Ğ°Ñ‡Ğ¸ÑĞ»ĞµĞ½Ğ¸Ğ¹ Ğ¿Ğ¾ SKU - ÑĞµĞ±ĞµÑÑ‚Ğ¾Ğ¸Ğ¼Ğ¾ÑÑ‚ÑŒ
Â  Â  Â  Â  Â  Â  Â  Â  by_sku_profit[sku] = by_sku_profit.get(sku, 0.0) + (amt - c)

Â  Â  Â  Â  Â  Â  net_profit = net_mp - cogs_total
Â  Â  Â  Â  Â  Â  margin = (net_profit / revenue * 100.0) if revenue > 0 else 0.0
Â  Â  Â  Â  Â  Â  status = "ğŸ”´" if net_profit <= 0 else ("ğŸŸ¡" if margin < 15 else "ğŸŸ¢")

Â  Â  Â  Â  Â  Â  msg = (
Â  Â  Â  Â  Â  Â  Â  Â  "ğŸŸ¢ Ğ§Ğ¸ÑÑ‚Ğ°Ñ Ğ¿Ñ€Ğ¸Ğ±Ñ‹Ğ»ÑŒ Ğ¿Ğ¾ Ğ¾Ñ‚Ñ‡Ñ‘Ñ‚Ñƒ OZON Â«ĞĞ°Ñ‡Ğ¸ÑĞ»ĞµĞ½Ğ¸ÑÂ»\n\n"
Â  Â  Â  Â  Â  Â  Â  Â  f"Ğ˜Ñ‚Ğ¾Ğ³Ğ¾ Ğ¾Ñ‚ OZON: {money(net_mp)}\n"
Â  Â  Â  Â  Â  Â  Â  Â  f"Ğ¡ĞµĞ±ĞµÑÑ‚Ğ¾Ğ¸Ğ¼Ğ¾ÑÑ‚ÑŒ: {money(cogs_total)}\n\n"
Â  Â  Â  Â  Â  Â  Â  Â  f"Ğ§Ğ¸ÑÑ‚Ğ°Ñ Ğ¿Ñ€Ğ¸Ğ±Ñ‹Ğ»ÑŒ: {money(net_profit)}\n"
Â  Â  Â  Â  Â  Â  Â  Â  f"ĞœĞ°Ñ€Ğ¶Ğ°: {pct(margin)}\n"
Â  Â  Â  Â  Â  Â  Â  Â  f"Ğ¡Ñ‚Ğ°Ñ‚ÑƒÑ: {status}\n\n"
Â  Â  Â  Â  Â  Â  Â  Â  f"Ğ¢ĞµÑ….Ğ¸Ğ½Ñ„Ğ¾: {parsed['note']}\n\n"
Â  Â  Â  Â  Â  Â  Â  Â  "Ğ¢ĞĞŸ-5 SKU Ğ¿Ğ¾ Ğ¿Ñ€Ğ¸Ğ±Ñ‹Ğ»Ğ¸:\n"
Â  Â  Â  Â  Â  Â  Â  Â  f"{top_lines_dict(by_sku_profit, 5, ascending=False)}\n\n"
Â  Â  Â  Â  Â  Â  Â  Â  "Ğ¢ĞĞŸ-5 SKU Ğ² Ğ¼Ğ¸Ğ½ÑƒÑ:\n"
Â  Â  Â  Â  Â  Â  Â  Â  f"{top_lines_dict(by_sku_profit, 5, ascending=True)}\n"
Â  Â  Â  Â  Â  Â  )

Â  Â  Â  Â  Â  Â  save_report(tg_id, {
Â  Â  Â  Â  Â  Â  Â  Â  "mode": "net_profit",
Â  Â  Â  Â  Â  Â  Â  Â  "file_name": file_name,
Â  Â  Â  Â  Â  Â  Â  Â  "revenue": revenue,
Â  Â  Â  Â  Â  Â  Â  Â  "deductions": deductions,
Â  Â  Â  Â  Â  Â  Â  Â  "net_mp": net_mp,
Â  Â  Â  Â  Â  Â  Â  Â  "cogs_total": cogs_total,
Â  Â  Â  Â  Â  Â  Â  Â  "net_profit": net_profit,
Â  Â  Â  Â  Â  Â  Â  Â  "margin": margin,
Â  Â  Â  Â  Â  Â  Â  Â  "note": parsed["note"],
Â  Â  Â  Â  Â  Â  })

Â  Â  Â  Â  Â  Â  context.user_data.clear()
Â  Â  Â  Â  Â  Â  await update.message.reply_text(msg, reply_markup=MAIN_KB)

Â  Â  Â  Â  except Exception as e:
Â  Â  Â  Â  Â  Â  await update.message.reply_text(
Â  Â  Â  Â  Â  Â  Â  Â  f"ĞĞµ ÑĞ¼Ğ¾Ğ³ Ñ€Ğ°Ğ·Ğ¾Ğ±Ñ€Ğ°Ñ‚ÑŒ Ñ„Ğ°Ğ¹Ğ» ğŸ˜•\n\nĞÑˆĞ¸Ğ±ĞºĞ°: {e}\n\n"
Â  Â  Â  Â  Â  Â  Â  Â  "ĞŸÑ€Ğ¾Ğ²ĞµÑ€ÑŒ, Ñ‡Ñ‚Ğ¾ ÑÑ‚Ğ¾ Excel (.xlsx) Ğ¸ ÑÑ‚Ğ¾ Ğ¾Ñ‚Ñ‡Ñ‘Ñ‚ OZON Â«ĞĞ°Ñ‡Ğ¸ÑĞ»ĞµĞ½Ğ¸ÑÂ».",
Â  Â  Â  Â  Â  Â  Â  Â  reply_markup=MODE_KB
Â  Â  Â  Â  Â  Â  )
Â  Â  Â  Â  return

Â  Â  await update.message.reply_text("Ğ¯ ÑĞµĞ¹Ñ‡Ğ°Ñ Ğ½Ğµ Ğ¶Ğ´Ñƒ Ñ„Ğ°Ğ¹Ğ». ĞĞ°Ğ¶Ğ¼Ğ¸ Â«ğŸ“ˆ ĞŸÑ€Ğ¸Ğ±Ñ‹Ğ»ÑŒ Ğ·Ğ° Ğ¿ĞµÑ€Ğ¸Ğ¾Ğ´Â».", reply_markup=MAIN_KB)


def main():
Â  Â  init_db()
Â  Â  app = Application.builder().token(TELEGRAM_BOT_TOKEN).build()
Â  Â  app.add_handler(CommandHandler("start", start))
Â  Â  app.add_handler(MessageHandler(filters.Document.ALL, handle_document))
Â  Â  app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))
Â  Â  app.run_polling()


if __name__ == "__main__":
Â  Â  main()
